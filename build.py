# 제주맛집.xlsx -> index.html 데이터 갱신 스크립트
# 사용법: python build.py
#   엑셀(Sheet1: 상호 / 위치·주소 / 대표메뉴·특징)을 읽어
#   index.html 안의 RESTAURANTS 배열만 새로 채워 넣는다.
#   (지도 코드/키/디자인은 그대로 보존)
import json, io, re, sys

XLSX = r'C:\Users\minsu\Desktop\제주맛집.xlsx'
HTML = r'C:\Users\minsu\workspace\jeju-map\index.html'

def region(addr):
    if '서귀포시' in addr: return '서귀포시'
    if '제주시' in addr: return '제주시'
    return '제주'

def category(menu):
    m = menu
    cafe   = ['빵','커피','카페','젤라또','케이크','크로플','도넛','타르트','까눌레','앙버터','크루아상','당근','땅콩크림','무스','라떼','치아바타']
    world  = ['양고기','양꼬치','할랄','인도','커리','파스타','이태리','스시','오마카세','이자까야','사시미','피자','맥주']
    noodle = ['국수','칼국수','냉면','짬뽕','국밥','감자탕','해장국','만두','떡볶이','죽','덮밥']
    sea    = ['갈치','고등어','회','물회','방어','생선','돔','딱새우','해물','전복','보말','오분자기','각재기','멜','문어','동태','삼합']
    meat   = ['흑돼지','돼지','삼겹','목살','돔베','오겹','소고기','정육','꼬치']
    for kw in cafe:
        if kw in m: return '카페·베이커리'
    for kw in world:
        if kw in m: return '세계요리'
    for kw in noodle:
        if kw in m: return '면·국물'
    for kw in sea:
        if kw in m: return '해산물·회'
    for kw in meat:
        if kw in m: return '흑돼지·고기'
    return '기타'

def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl 필요: pip install openpyxl')

    wb = openpyxl.load_workbook(XLSX)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]

    data = []
    for r in rows[1:]:  # 헤더 제외
        name = (r[0] or '').strip() if r and r[0] else ''
        addr = (r[1] or '').strip() if len(r) > 1 and r[1] else ''
        menu = (r[2] or '').strip() if len(r) > 2 and r[2] else ''
        if not name:
            continue
        data.append({'name': name, 'addr': addr, 'menu': menu,
                     'region': region(addr), 'cat': category(menu)})

    data_js = json.dumps(data, ensure_ascii=False)

    html = io.open(HTML, encoding='utf-8').read()
    new_html, n = re.subn(r'const RESTAURANTS = .*?;\n',
                          'const RESTAURANTS = ' + data_js + ';\n',
                          html, count=1)
    if n != 1:
        sys.exit('index.html에서 RESTAURANTS 배열을 찾지 못했습니다. (구조 변경됨?)')

    io.open(HTML, 'w', encoding='utf-8').write(new_html)
    print('OK - %d개 맛집으로 index.html 갱신 완료' % len(data))

if __name__ == '__main__':
    main()
